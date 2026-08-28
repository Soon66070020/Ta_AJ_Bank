"""
process_shipment.py
-------------------
Processes and standardizes Industrial Shipment Index (ดัชนีการส่งสินค้า) Excel file
from presentation format (.xlsx) to clean, structured CSV format with UTF-8 BOM encoding.

Data Source: Office of Industrial Economics (สศอ.), Ministry of Industry
Classification Standard: TSIC (Thailand Standard Industrial Classification)
Base Year: 2559 = 100 (2016 = 100)

Author: Antigravity Assistant
Environment: Python 3.11+ / venv
"""

import os
import re
import sys
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import openpyxl
import pandas as pd
import numpy as np

# Configure UTF-8 stdout for Windows console
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s'
)

TH_MONTH_MAP = {
    'ม.ค.': 1, 'ก.พ.': 2, 'มี.ค.': 3, 'เม.ย.': 4,
    'พ.ค.': 5, 'มิ.ย.': 6, 'มิ.ย.*': 6,
    'ก.ค.': 7, 'ส.ค.': 8, 'ก.ย.': 9, 'ต.ค.': 10,
    'พ.ย.': 11, 'ธ.ค.': 12
}

TH_MONTH_NAMES = {
    1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน',
    5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฎาคม', 8: 'สิงหาคม',
    9: 'กันยายน', 10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'
}


def clean_num(val: Any) -> Optional[float]:
    """Converts cell values to float or None."""
    if val is None or val == "" or str(val).strip() in ["-", "N/A", "n/a", "null", ""]:
        return None
    try:
        if isinstance(val, (int, float)):
            return round(float(val), 6)
        s = str(val).replace(',', '').strip()
        return round(float(s), 6)
    except (ValueError, TypeError):
        return None


def parse_time_columns(ws: Any) -> Tuple[List[Dict[str, Any]], int, int]:
    """
    Parses Row 4 (Year) and Row 5 (Month) to construct canonical time column metadata.
    Returns: (time_cols_meta, col_mom_idx, col_yoy_idx)
    """
    time_cols = []
    current_year_be = None
    col_mom_idx = -1
    col_yoy_idx = -1

    for c in range(5, ws.max_column + 1):
        r4_val = ws.cell(4, c).value
        r5_val = ws.cell(5, c).value

        if r4_val:
            r4_str = str(r4_val).strip()
            if r4_str.isdigit():
                current_year_be = int(r4_str)
            elif '% Change' in r4_str and 'เดือนก่อน' in r4_str:
                col_mom_idx = c
                continue
            elif '% Change' in r4_str and 'ปีก่อน' in r4_str:
                col_yoy_idx = c
                continue

        if r5_val is not None and current_year_be:
            m_str = str(r5_val).strip()
            is_preliminary = '*' in m_str
            clean_m_str = m_str.replace('*', '')
            month_num = TH_MONTH_MAP.get(m_str) or TH_MONTH_MAP.get(clean_m_str)
            
            if month_num:
                year_ce = current_year_be - 543
                period_ym = f"{year_ce}-{month_num:02d}"
                col_name_wide = f"index_{year_ce}_{month_num:02d}"

                time_cols.append({
                    'col_idx': c,
                    'year_be': current_year_be,
                    'year_ce': year_ce,
                    'month_num': month_num,
                    'month_name_th': TH_MONTH_NAMES.get(month_num, ''),
                    'period_ym': period_ym,
                    'col_name_wide': col_name_wide,
                    'is_preliminary': is_preliminary
                })

    return time_cols, col_mom_idx, col_yoy_idx


def process_shipment_file(
    input_file: str = 'group_5/raw/Shipment.xlsx',
    output_dir: str = 'group_5/clean_csv'
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Transforms Shipment.xlsx into:
    1. shipment_index_wide.csv
    2. shipment_index_tidy_timeseries.csv
    """
    input_path = Path(input_file)
    if not input_path.exists() and Path('group_5/Shipment.xlsx').exists():
        input_path = Path('group_5/Shipment.xlsx')

    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    ws = wb['Shipment']

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    base_year_info = "2559 = 100"
    source_agency = "สำนักงานเศรษฐกิจอุตสาหกรรม (สศอ.) กระทรวงอุตสาหกรรม"
    source_filename = input_path.name

    # 1. Parse Column Definitions
    time_cols, col_mom_idx, col_yoy_idx = parse_time_columns(ws)
    logging.info(f"Identified {len(time_cols)} monthly time columns ({time_cols[0]['period_ym']} to {time_cols[-1]['period_ym']})")

    # 2. Parse Rows & TSIC Hierarchy
    wide_records = []
    tidy_records = []

    current_div_code = ""
    current_div_name = ""
    current_grp_code = ""
    current_grp_name = ""
    current_cls_code = ""
    current_cls_name = ""

    item_counter = 0

    for r in range(6, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value is not None else ""
        c2 = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value is not None else ""
        c3 = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value is not None else ""
        weight_val = clean_num(ws.cell(r, 4).value)

        if not c3:
            continue

        item_counter += 1

        # TSIC Hierarchy Classifier
        if c3 == 'ดัชนีรวมยังไม่ได้ปรับฤดูกาล':
            cat_level = 0
            level_type = 'TOTAL'
            tsic_div_code = ""
            tsic_div_name = "ดัชนีรวมยังไม่ได้ปรับฤดูกาล"
            tsic_grp_code = ""
            tsic_grp_name = ""
            tsic_cls_code = ""
            tsic_cls_name = ""
            prod_code = ""
            item_name = c3
            current_div_code = ""
            current_div_name = ""
            current_grp_code = ""
            current_grp_name = ""
            current_cls_code = ""
            current_cls_name = ""
        elif c3.startswith('TSIC :') and len(c3.split()[2]) == 2:
            cat_level = 1
            level_type = 'DIVISION_2DIGIT'
            parts = c3.split(maxsplit=3)
            current_div_code = parts[2]
            current_div_name = parts[3] if len(parts) > 3 else ""
            current_grp_code = ""
            current_grp_name = ""
            current_cls_code = ""
            current_cls_name = ""
            tsic_div_code = current_div_code
            tsic_div_name = current_div_name
            tsic_grp_code = ""
            tsic_grp_name = ""
            tsic_cls_code = ""
            tsic_cls_name = ""
            prod_code = ""
            item_name = current_div_name
        elif c3.startswith('TSIC :') and len(c3.split()[2]) == 4:
            cat_level = 2
            level_type = 'GROUP_4DIGIT'
            parts = c3.split(maxsplit=3)
            current_grp_code = parts[2]
            current_grp_name = parts[3] if len(parts) > 3 else ""
            current_cls_code = ""
            current_cls_name = ""
            tsic_div_code = current_div_code
            tsic_div_name = current_div_name
            tsic_grp_code = current_grp_code
            tsic_grp_name = current_grp_name
            tsic_cls_code = ""
            tsic_cls_name = ""
            prod_code = ""
            item_name = current_grp_name
        elif not c1 and not c2 and re.match(r'^\d{5}\s+', c3):
            cat_level = 3
            level_type = 'CLASS_5DIGIT'
            m = re.match(r'^(\d{5})\s+(.*)', c3)
            current_cls_code = m.group(1)
            current_cls_name = m.group(2).strip()
            tsic_div_code = current_div_code
            tsic_div_name = current_div_name
            tsic_grp_code = current_grp_code
            tsic_grp_name = current_grp_name
            tsic_cls_code = current_cls_code
            tsic_cls_name = current_cls_name
            prod_code = ""
            item_name = current_cls_name
        elif c1 and c2:
            cat_level = 4
            level_type = 'PRODUCT_ITEM'
            tsic_div_code = current_div_code
            tsic_div_name = current_div_name
            tsic_grp_code = current_grp_code
            tsic_grp_name = current_grp_name
            tsic_cls_code = current_cls_code
            tsic_cls_name = current_cls_name
            prod_code = f"{c1}-{c2}"
            item_name = c3
        else:
            cat_level = -1
            level_type = 'OTHER'
            tsic_div_code = current_div_code
            tsic_div_name = current_div_name
            tsic_grp_code = current_grp_code
            tsic_grp_name = current_grp_name
            tsic_cls_code = current_cls_code
            tsic_cls_name = current_cls_name
            prod_code = f"{c1}-{c2}" if c1 and c2 else ""
            item_name = c3

        mom_val = clean_num(ws.cell(r, col_mom_idx).value) if col_mom_idx > 0 else None
        yoy_val = clean_num(ws.cell(r, col_yoy_idx).value) if col_yoy_idx > 0 else None

        # Build Wide Record
        wide_row = {
            'row_order': item_counter,
            'category_level': cat_level,
            'level_type': level_type,
            'tsic_division_code': tsic_div_code,
            'tsic_division_name': tsic_div_name,
            'tsic_group_code': tsic_grp_code,
            'tsic_group_name': tsic_grp_name,
            'tsic_class_code': tsic_cls_code,
            'tsic_class_name': tsic_cls_name,
            'product_code': prod_code,
            'item_name': item_name,
            'weight': weight_val
        }

        # Extract monthly values for both Wide and Tidy
        for t in time_cols:
            cell_val = clean_num(ws.cell(r, t['col_idx']).value)
            wide_row[t['col_name_wide']] = cell_val

            tidy_records.append({
                'period_ym': t['period_ym'],
                'year_ce': t['year_ce'],
                'year_be': t['year_be'],
                'month_num': t['month_num'],
                'month_name_th': t['month_name_th'],
                'row_order': item_counter,
                'category_level': cat_level,
                'level_type': level_type,
                'tsic_division_code': tsic_div_code,
                'tsic_division_name': tsic_div_name,
                'tsic_group_code': tsic_grp_code,
                'tsic_group_name': tsic_grp_name,
                'tsic_class_code': tsic_cls_code,
                'tsic_class_name': tsic_cls_name,
                'product_code': prod_code,
                'item_name': item_name,
                'weight': weight_val,
                'shipment_index': cell_val,
                'is_preliminary': t['is_preliminary'],
                'base_year': base_year_info,
                'source_agency': source_agency,
                'source_file': source_filename
            })

        wide_row['mom_change_pct'] = mom_val
        wide_row['yoy_change_pct'] = yoy_val
        wide_row['base_year'] = base_year_info
        wide_row['source_agency'] = source_agency
        wide_row['source_file'] = source_filename

        wide_records.append(wide_row)

    df_wide = pd.DataFrame(wide_records)
    df_tidy = pd.DataFrame(tidy_records)

    # Save Wide CSV
    wide_csv_path = output_path / "shipment_index_wide.csv"
    df_wide.to_csv(wide_csv_path, index=False, encoding='utf-8-sig')
    logging.info(f"Saved Wide Format CSV: {wide_csv_path} ({len(df_wide)} rows x {len(df_wide.columns)} cols)")

    # Save Tidy Time-Series CSV
    tidy_csv_path = output_path / "shipment_index_tidy_timeseries.csv"
    df_tidy.to_csv(tidy_csv_path, index=False, encoding='utf-8-sig')
    logging.info(f"Saved Tidy Time-Series CSV: {tidy_csv_path} ({len(df_tidy)} rows x {len(df_tidy.columns)} cols)")

    return df_wide, df_tidy


if __name__ == '__main__':
    in_file = sys.argv[1] if len(sys.argv) > 1 else 'group_5/raw/Shipment.xlsx'
    out_dir = sys.argv[2] if len(sys.argv) > 2 else 'group_5/clean_csv'
    process_shipment_file(input_file=in_file, output_dir=out_dir)
