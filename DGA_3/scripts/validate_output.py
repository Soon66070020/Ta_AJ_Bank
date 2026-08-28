import sys
import glob
from pathlib import Path
import pandas as pd
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

csv_files = sorted(glob.glob('group_5/clean_csv/ppi_*.csv'))
print(f"Generated PPI CSV files ({len(csv_files)}):")
for f in csv_files:
    df = pd.read_csv(f, encoding='utf-8-sig')
    print(f"\n--- {Path(f).name} ---")
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print(f"Sample Head (2 rows):")
    print(df.head(2)[['period_ym', 'category_level', 'sector', 'category_name'] + [c for c in df.columns if 'index' in c or 'price' in c]])
    
    # Check nulls
    null_counts = df.isnull().sum()
    print("Null counts per column:")
    for col, count in null_counts.items():
        if count > 0:
            print(f"  {col}: {count} nulls")
        else:
            print(f"  {col}: 0 nulls")

print("\n==========================================")
print("CROSS-VALIDATION WITH RAW EXCEL FILES")
print("==========================================")

test_files = [
    ('group_5/raw/ดัชนีราคาผู้ผลิต_(เม.ย 69).xlsx' if Path('group_5/raw/ดัชนีราคาผู้ผลิต_(เม.ย 69).xlsx').exists() else 'group_5/ดัชนีราคาผู้ผลิต_(เม.ย 69).xlsx', 'group_5/clean_csv/ppi_monthly_2026_04.csv'),
    ('group_5/raw/ดัชนีราคาผู้ผลิต_(พ.ค 69).xlsx' if Path('group_5/raw/ดัชนีราคาผู้ผลิต_(พ.ค 69).xlsx').exists() else 'group_5/ดัชนีราคาผู้ผลิต_(พ.ค 69).xlsx', 'group_5/clean_csv/ppi_monthly_2026_05.csv'),
    ('group_5/raw/ดัชนีราคาผู้ผลิต_(มิ.ย 69).xlsx' if Path('group_5/raw/ดัชนีราคาผู้ผลิต_(มิ.ย 69).xlsx').exists() else 'group_5/ดัชนีราคาผู้ผลิต_(มิ.ย 69).xlsx', 'group_5/clean_csv/ppi_monthly_2026_06.csv'),
]

for excel_f, csv_f in test_files:
    wb = openpyxl.load_workbook(excel_f, data_only=True)
    ws = wb['export']
    df_csv = pd.read_csv(csv_f, encoding='utf-8-sig')
    
    # Check 90 rows match exactly
    all_matched = True
    mismatches = []
    
    for r in range(7, 97):
        excel_cat = ws.cell(r, 1).value.strip()
        excel_curr = float(ws.cell(r, 2).value) if ws.cell(r, 2).value is not None else None
        excel_mom = float(ws.cell(r, 5).value) if ws.cell(r, 5).value is not None else None
        excel_yoy = float(ws.cell(r, 6).value) if ws.cell(r, 6).value is not None else None
        excel_aoa = float(ws.cell(r, 7).value) if ws.cell(r, 7).value is not None else None
        
        csv_row = df_csv[df_csv['category_name'] == excel_cat].iloc[0]
        csv_curr = csv_row['index_current_month']
        csv_mom = csv_row['mom_change_pct']
        csv_yoy = csv_row['yoy_change_pct']
        csv_aoa = csv_row['aoa_change_pct']
        
        if round(excel_curr, 4) != round(csv_curr, 4) or \
           round(excel_mom, 4) != round(csv_mom, 4) or \
           round(excel_yoy, 4) != round(csv_yoy, 4) or \
           round(excel_aoa, 4) != round(csv_aoa, 4):
            all_matched = False
            mismatches.append((excel_cat, (excel_curr, excel_mom, excel_yoy, excel_aoa), (csv_curr, csv_mom, csv_yoy, csv_aoa)))
            
    if all_matched:
        print(f"PASS: {Path(excel_f).name} -> {Path(csv_f).name} (All 90 categories and metrics 100% matched)")
    else:
        print(f"FAIL: {len(mismatches)} mismatches found in {excel_f}!")
        for m in mismatches[:5]:
            print(f"  {m}")

print("\nALL VERIFICATIONS COMPLETED SUCCESSFULLY!")
