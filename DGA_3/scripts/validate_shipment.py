import sys
from pathlib import Path
import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')

print("==========================================")
print("VALIDATING SHIPMENT CSV DATASETS")
print("==========================================")

wide_path = 'group_5/clean_csv/shipment_index_wide.csv'
tidy_path = 'group_5/clean_csv/shipment_index_tidy_timeseries.csv'

df_wide = pd.read_csv(wide_path, encoding='utf-8-sig')
df_tidy = pd.read_csv(tidy_path, encoding='utf-8-sig')

print(f"Wide CSV shape: {df_wide.shape} (Expected: 523 rows)")
print(f"Tidy CSV shape: {df_tidy.shape} (Expected: 34518 rows = 523 x 66)")

assert df_wide.shape[0] == 523, "Wide row count mismatch!"
assert df_tidy.shape[0] == 34518, "Tidy row count mismatch!"

print("Level distribution in Wide CSV:")
print(df_wide['level_type'].value_counts())

print("\nSample Tidy Records:")
print(df_tidy[['period_ym', 'level_type', 'tsic_division_name', 'product_code', 'item_name', 'shipment_index']].head(5))

# Cross-Validation with raw Shipment.xlsx
raw_shipment_file = 'group_5/raw/Shipment.xlsx' if Path('group_5/raw/Shipment.xlsx').exists() else 'group_5/Shipment.xlsx'
wb = openpyxl.load_workbook(raw_shipment_file, data_only=True)
ws = wb['Shipment']

# Test specific items
sample_test_rows = [6, 7, 10, 16, 24, 100, 250, 400, 524, 528]

for r in sample_test_rows:
    raw_item_name = ws.cell(r, 3).value
    raw_weight = float(ws.cell(r, 4).value)
    raw_val_2021_01 = float(ws.cell(r, 5).value)
    raw_val_2026_06 = float(ws.cell(r, 70).value)
    raw_mom = float(ws.cell(r, 71).value)
    raw_yoy = float(ws.cell(r, 72).value)

    item_order = r - 5
    wide_row = df_wide[df_wide['row_order'] == item_order].iloc[0]

    assert round(wide_row['weight'], 4) == round(raw_weight, 4), f"Weight mismatch at row {r}"
    assert round(wide_row['index_2021_01'], 4) == round(raw_val_2021_01, 4), f"2021_01 mismatch at row {r}"
    assert round(wide_row['index_2026_06'], 4) == round(raw_val_2026_06, 4), f"2026_06 mismatch at row {r}"
    assert round(wide_row['mom_change_pct'], 4) == round(raw_mom, 4), f"MoM mismatch at row {r}"
    assert round(wide_row['yoy_change_pct'], 4) == round(raw_yoy, 4), f"YoY mismatch at row {r}"

    # Check in tidy
    tidy_jun_26 = df_tidy[(df_tidy['row_order'] == item_order) & (df_tidy['period_ym'] == '2026-06')].iloc[0]
    assert round(tidy_jun_26['shipment_index'], 4) == round(raw_val_2026_06, 4), f"Tidy mismatch at row {r}"
    assert tidy_jun_26['is_preliminary'] == True, f"Preliminary flag error at row {r}"

    print(f"PASS: Row {r:3d} -> '{wide_row['item_name']}' ({wide_row['level_type']}) - All values matched 100%!")

print("\nALL SHIPMENT VERIFICATIONS COMPLETED SUCCESSFULLY!")
