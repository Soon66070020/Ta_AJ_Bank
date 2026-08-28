"""
process_price_index.py
----------------------
Processes and standardizes Thai Producer Price Index (ดัชนีราคาผู้ผลิต - PPI) Excel files
from presentation format (.xlsx) to clean, structured CSV format with UTF-8 BOM encoding.

Author: Antigravity Assistant
Environment: Python 3.11+ / venv
"""

import os
import re
import sys
import glob
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import openpyxl
import pandas as pd
import numpy as np

# Configure UTF-8 stdout for Windows console
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s'
)

# Thai Month Mapping (Name / Abbreviation to Month Number 1-12)
TH_MONTHS = {
    'มกราคม': 1, 'ม.ค.': 1, 'ม.ค': 1,
    'กุมภาพันธ์': 2, 'ก.พ.': 2, 'ก.พ': 2,
    'มีนาคม': 3, 'มี.ค.': 3, 'มี.ค': 3,
    'เมษายน': 4, 'เม.ย.': 4, 'เม.ย': 4,
    'พฤษภาคม': 5, 'พ.ค.': 5, 'พ.ค': 5,
    'มิถุนายน': 6, 'มิ.ย.': 6, 'มิ.ย': 6,
    'กรกฎาคม': 7, 'ก.ค.': 7, 'ก.ค': 7,
    'สิงหาคม': 8, 'ส.ค.': 8, 'ส.ค': 8,
    'กันยายน': 9, 'ก.ย.': 9, 'ก.ย': 9,
    'ตุลาคม': 10, 'ต.ค.': 10, 'ต.ค': 10,
    'พฤศจิกายน': 11, 'พ.ย.': 11, 'พ.ย': 11,
    'ธันวาคม': 12, 'ธ.ค.': 12, 'ธ.ค': 12
}

# Standard CPA Hierarchy Definitions
SECTORS_L1 = {
    'ผลิตภัณฑ์เกษตรกรรม และการประมง',
    'ผลิตภัณฑ์จากเหมือง',
    'ผลิตภัณฑ์อุตสาหกรรม'
}

GROUPS_L2_AGRICULTURE = {
    'ผลิตภัณฑ์ทางการเกษตร',
    'ผลิตภัณฑ์จากการประมง'
}

GROUPS_L2_MINING = {
    'ถ่านหินและลิกไนต์',
    'ปิโตรเลียมดิบและก๊าซธรรมชาติ',
    'สินแร่โลหะ',
    'ผลิตภัณฑ์อื่นๆ ที่ได้จากการทำเหมือง'
}

GROUPS_L2_MANUFACTURING = {
    'ผลิตภัณฑ์อาหาร',
    'เครื่องดื่ม',
    'ผลิตภัณฑ์ยาสูบ',
    'สิ่งทอ',
    'เสื้อผ้าเครื่องแต่งกาย',
    'เครื่องหนังและผลิตภัณฑ์',
    'ไม้และผลิตภัณฑ์จากไม้',
    'กระดาษและผลิตภัณฑ์กระดาษ',
    'สิ่งพิมพ์และสิ่งที่เกี่ยวข้องกับการพิมพ์',
    'ผลิตภัณฑ์ที่ได้จากการกลั่นปิโตรเลียม',
    'เคมีภัณฑ์และผลิตภัณฑ์เคมี',
    'ผลิตภัณฑ์ทางเภสัชกรรม',
    'ผลิตภัณฑ์ยางและพลาสติก',
    'ผลิตภัณฑ์ที่ทำจากแร่อโลหะ',
    'โลหะขั้นมูลฐาน',
    'ผลิตภัณฑ์โลหะประดิษฐ์',
    'ผลิตภัณฑ์คอมพิวเตอร์ และอิเล็กทรอนิกส์',
    'อุปกรณ์ไฟฟ้า',
    'เครื่องจักรและเครื่องมือ',
    'ยานยนต์ ชิ้นส่วนและอุปกรณ์',
    'จักรยานยนต์ จักรยาน ชิ้นส่วนและอุปกรณ์',
    'อุปกรณ์ขนส่ง ซึ่งมิได้จัดประเภทไว้ในที่อื่น',
    'เฟอร์นิเจอร์',
    'ผลิตภัณฑ์อุตสาหกรรมอื่นๆ'
}


def parse_thai_period(text: str) -> Dict[str, Any]:
    """
    Parses Thai period string like 'ระยะเวลา พฤษภาคม 2569' into structured date metadata.
    """
    if not text:
        return {}
    clean_text = str(text).strip()
    match = re.search(r'([ก-๙\.]+)\s+(\d{4})', clean_text)
    if match:
        month_str = match.group(1).strip()
        year_be = int(match.group(2))
        month_num = TH_MONTHS.get(month_str, None)
        year_ce = year_be - 543
        period_ym = f"{year_ce}-{month_num:02d}" if month_num else None
        return {
            'period_raw': clean_text,
            'month_name_th': month_str,
            'month_num': month_num,
            'year_be': year_be,
            'year_ce': year_ce,
            'period_ym': period_ym
        }
    return {'period_raw': clean_text}


def parse_base_year(text: str) -> str:
    """Extracts base year text, e.g. 'ปีฐาน (2564 = 100)' -> '2564 = 100'."""
    if not text:
        return ""
    m = re.search(r'\((.*?)\)', str(text))
    if m:
        return m.group(1).strip()
    return str(text).replace('ปีฐาน', '').strip()


def clean_numeric(val: Any) -> Optional[float]:
    """Converts Excel cell values to clean float or None."""
    if val is None or val == "" or str(val).strip() in ["-", "N/A", "n/a", "null", ""]:
        return None
    try:
        if isinstance(val, (int, float)):
            return round(float(val), 4)
        s = str(val).replace(',', '').strip()
        return round(float(s), 4)
    except (ValueError, TypeError):
        return None


def classify_cpa_hierarchy(categories: List[str]) -> List[Dict[str, Any]]:
    """
    Walks through the sequential 90 CPA categories to identify hierarchy level (0..3),
    current parent sector (L1), and current group (L2).
    """
    results = []
    current_sector = None
    current_group = None

    for idx, raw_cat in enumerate(categories, 1):
        cat = raw_cat.strip()
        
        if cat == 'รวมทุกรายการ':
            level = 0
            sector = 'รวมทุกรายการ'
            group = 'รวมทุกรายการ'
            current_sector = None
            current_group = None
        elif cat in SECTORS_L1:
            level = 1
            current_sector = cat
            current_group = None
            sector = current_sector
            group = None
        elif current_sector == 'ผลิตภัณฑ์เกษตรกรรม และการประมง' and cat in GROUPS_L2_AGRICULTURE:
            level = 2
            current_group = cat
            sector = current_sector
            group = current_group
        elif current_sector == 'ผลิตภัณฑ์จากเหมือง' and cat in GROUPS_L2_MINING:
            level = 2
            current_group = cat
            sector = current_sector
            group = current_group
        elif current_sector == 'ผลิตภัณฑ์อุตสาหกรรม' and cat in GROUPS_L2_MANUFACTURING:
            level = 2
            current_group = cat
            sector = current_sector
            group = current_group
        else:
            # Sub-category (Level 3)
            level = 3
            sector = current_sector
            group = current_group

        results.append({
            'row_order': idx,
            'category_name': cat,
            'category_level': level,
            'sector': sector,
            'group_name': group
        })

    return results


def process_single_ppi_file(filepath: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Reads a single PPI Excel file and returns a structured DataFrame and its metadata.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = wb.sheetnames[0]  # 'export'
    ws = wb[sheet_name]

    # 1. Read Metadata from Rows 1-4
    r1_val = ws.cell(1, 1).value
    r2_val = ws.cell(2, 1).value
    r3_val = ws.cell(3, 1).value
    r4_val = ws.cell(4, 1).value

    period_meta = parse_thai_period(r1_val)
    base_year_meta = parse_base_year(r2_val)
    view_meta = str(r3_val).strip() if r3_val else ""
    dl_meta = str(r4_val).strip() if r4_val else ""

    # 2. Extract Category names from Row 7 onwards
    raw_rows = []
    category_names = []
    for r in range(7, ws.max_row + 1):
        cat_val = ws.cell(r, 1).value
        if cat_val is not None and str(cat_val).strip():
            cat_name = str(cat_val).strip()
            category_names.append(cat_name)
            
            # Numeric columns B..G
            idx_curr = clean_numeric(ws.cell(r, 2).value)
            idx_prev_m = clean_numeric(ws.cell(r, 3).value)
            idx_prev_y = clean_numeric(ws.cell(r, 4).value)
            mom = clean_numeric(ws.cell(r, 5).value)
            yoy = clean_numeric(ws.cell(r, 6).value)
            aoa = clean_numeric(ws.cell(r, 7).value)

            raw_rows.append({
                'category_name': cat_name,
                'index_current_month': idx_curr,
                'index_prev_month': idx_prev_m,
                'index_prev_year_same_month': idx_prev_y,
                'mom_change_pct': mom,
                'yoy_change_pct': yoy,
                'aoa_change_pct': aoa
            })

    # 3. Classify Hierarchy
    hierarchy_info = classify_cpa_hierarchy(category_names)
    
    # 4. Construct Clean DataFrame
    df_rows = []
    filename = Path(filepath).name
    for h, data in zip(hierarchy_info, raw_rows):
        df_rows.append({
            'period_ym': period_meta.get('period_ym'),
            'year_ce': period_meta.get('year_ce'),
            'year_be': period_meta.get('year_be'),
            'month_num': period_meta.get('month_num'),
            'month_name_th': period_meta.get('month_name_th'),
            'row_order': h['row_order'],
            'category_level': h['category_level'],
            'sector': h['sector'],
            'group_name': h['group_name'],
            'category_name': h['category_name'],
            'index_current_month': data['index_current_month'],
            'index_prev_month': data['index_prev_month'],
            'index_prev_year_same_month': data['index_prev_year_same_month'],
            'mom_change_pct': data['mom_change_pct'],
            'yoy_change_pct': data['yoy_change_pct'],
            'aoa_change_pct': data['aoa_change_pct'],
            'base_year': base_year_meta,
            'source_file': filename
        })

    df = pd.DataFrame(df_rows)
    metadata = {
        'filepath': filepath,
        'filename': filename,
        'period_meta': period_meta,
        'base_year': base_year_meta,
        'view_title': view_meta,
        'download_info': dl_meta,
        'num_records': len(df)
    }
    return df, metadata


def process_all_ppi_files(input_dir: str = 'group_5/raw', output_dir: str = 'group_5/clean_csv') -> None:
    """
    Processes all PPI Excel files in `input_dir` and saves cleaned CSVs into `output_dir`.
    """
    input_path = Path(input_dir)
    if not input_path.exists() and Path('group_5').exists():
        input_path = Path('group_5')

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    pattern = str(input_path / '*ดัชนีราคา*.xlsx')
    excel_files = sorted(glob.glob(pattern))

    if not excel_files:
        logging.warning(f"No files matching '{pattern}' found.")
        return

    logging.info(f"Found {len(excel_files)} PPI files in {input_dir}")

    all_monthly_dfs = []

    for f in excel_files:
        logging.info(f"Processing: {f}")
        df, meta = process_single_ppi_file(f)
        all_monthly_dfs.append(df)

        # Save monthly clean CSV
        period_ym = meta['period_meta'].get('period_ym', 'unknown')
        monthly_csv_name = f"ppi_monthly_{period_ym.replace('-', '_')}.csv"
        monthly_csv_path = output_path / monthly_csv_name
        
        # Save with utf-8-sig for perfect Excel / Thai encoding compatibility
        df.to_csv(monthly_csv_path, index=False, encoding='utf-8-sig')
        logging.info(f"  -> Saved monthly snapshot: {monthly_csv_path} ({len(df)} rows)")

    # Consolidate into Master Time-Series dataset
    if all_monthly_dfs:
        master_df = pd.concat(all_monthly_dfs, ignore_index=True)
        # Sort chronologically and by row order
        master_df = master_df.sort_values(by=['period_ym', 'row_order']).reset_index(drop=True)

        master_csv_path = output_path / "ppi_timeseries_master.csv"
        master_df.to_csv(master_csv_path, index=False, encoding='utf-8-sig')
        logging.info(f"Consolidated Master Time-Series CSV: {master_csv_path} ({len(master_df)} rows total)")

        # Create Tidy Long-Format Dataset (Single observation per row: period, category, metric, value)
        tidy_records = []
        for _, row in master_df.iterrows():
            tidy_records.append({
                'period_ym': row['period_ym'],
                'year_ce': row['year_ce'],
                'month_num': row['month_num'],
                'category_level': row['category_level'],
                'sector': row['sector'],
                'group_name': row['group_name'],
                'category_name': row['category_name'],
                'price_index': row['index_current_month'],
                'mom_pct': row['mom_change_pct'],
                'yoy_pct': row['yoy_change_pct'],
                'aoa_pct': row['aoa_change_pct'],
                'base_year': row['base_year'],
                'source_file': row['source_file']
            })
        
        tidy_df = pd.DataFrame(tidy_records)
        tidy_csv_path = output_path / "ppi_tidy_timeseries.csv"
        tidy_df.to_csv(tidy_csv_path, index=False, encoding='utf-8-sig')
        logging.info(f"Tidy Time-Series CSV: {tidy_csv_path} ({len(tidy_df)} rows total)")


if __name__ == '__main__':
    in_dir = sys.argv[1] if len(sys.argv) > 1 else 'group_5/raw'
    out_dir = sys.argv[2] if len(sys.argv) > 2 else 'group_5/clean_csv'
    process_all_ppi_files(input_dir=in_dir, output_dir=out_dir)
